import pandas as pd
import PyPDF2
import pytesseract
from PIL import Image
import openpyxl
import csv
import json
import re
from typing import Dict, Any, List
from pathlib import Path

class FileProcessor:
    """Handles file upload, parsing, and data extraction"""
    
    def __init__(self):
        self.supported_formats = {
            'pdf': self._process_pdf,
            'xlsx': self._process_excel,
            'xls': self._process_excel,
            'csv': self._process_csv,
            'png': self._process_image_ocr,
            'jpg': self._process_image_ocr,
            'jpeg': self._process_image_ocr
        }
    
    def detect_file_type(self, filename: str) -> str:
        """Detect file type from extension"""
        ext = Path(filename).suffix.lower().lstrip('.')
        if ext in self.supported_formats:
            return ext
        raise ValueError(f"Unsupported file type: {ext}")
    
    def process_file(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """Main processing router"""
        processor = self.supported_formats.get(file_type)
        if not processor:
            raise ValueError(f"No processor for file type: {file_type}")
        
        raw_data = processor(file_path)
        structured_data = self._structure_financial_data(raw_data)
        return structured_data
    
    def _process_pdf(self, file_path: str) -> Dict[str, Any]:
        """Extract data from PDF"""
        text_content = []
        tables = []
        
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text = page.extract_text()
                text_content.append(text)
        
        full_text = "\n".join(text_content)
        
        # Try to extract tabular data using pattern matching
        tables = self._extract_tables_from_text(full_text)
        
        return {
            "text": full_text,
            "tables": tables,
            "page_count": len(text_content)
        }
    
    def _process_excel(self, file_path: str) -> Dict[str, Any]:
        """Extract data from Excel"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_data = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            sheets_data[sheet_name] = data
        
        # Convert to pandas for easier processing
        dfs = {}
        for sheet_name, data in sheets_data.items():
            if data:
                df = pd.DataFrame(data[1:], columns=data[0])
                dfs[sheet_name] = df.to_dict('records')
        
        return {
            "sheets": dfs,
            "sheet_names": list(dfs.keys())
        }
    
    def _process_csv(self, file_path: str) -> Dict[str, Any]:
        """Extract data from CSV"""
        df = pd.read_csv(file_path)
        return {
            "data": df.to_dict('records'),
            "columns": list(df.columns),
            "rows": len(df)
        }
    
    def _process_image_ocr(self, file_path: str) -> Dict[str, Any]:
        """Extract text from image using OCR"""
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        
        tables = self._extract_tables_from_text(text)
        
        return {
            "text": text,
            "tables": tables
        }
    
    def _extract_tables_from_text(self, text: str) -> List[Dict]:
        """Extract tabular data from text using pattern matching"""
        # Basic table detection - looks for aligned numeric data
        tables = []
        lines = text.split('\n')
        
        current_table = []
        for line in lines:
            # Check if line contains multiple numbers (likely a data row)
            numbers = re.findall(r'-?\d+(?:\.\d+)?', line)
            if len(numbers) >= 2:
                current_table.append(line)
            elif current_table:
                if len(current_table) >= 2:
                    tables.append({"rows": current_table})
                current_table = []
        
        return tables
    
    def _structure_financial_data(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """Convert raw extracted data into structured financial format"""
        
        structured = {
            "raw": raw_data,
            "balance_sheet": {},
            "income_statement": {},
            "cash_flow": {},
            "ratios": {},
            "metadata": {}
        }
        
        # Attempt to identify financial statement sections
        if "text" in raw_data:
            text = raw_data["text"].lower()
            
            # Extract key financial figures using pattern matching
            structured["balance_sheet"] = self._extract_balance_sheet(text)
            structured["income_statement"] = self._extract_income_statement(text)
            structured["cash_flow"] = self._extract_cash_flow(text)
        
        if "sheets" in raw_data:
            # Excel file - try to map sheets to statements
            for sheet_name, data in raw_data["sheets"].items():
                sheet_lower = sheet_name.lower()
                if "balance" in sheet_lower or "bs" in sheet_lower:
                    structured["balance_sheet"] = data
                elif "income" in sheet_lower or "p&l" in sheet_lower or "pnl" in sheet_lower:
                    structured["income_statement"] = data
                elif "cash" in sheet_lower or "cf" in sheet_lower:
                    structured["cash_flow"] = data
        
        if "data" in raw_data:
            # CSV file
            structured["data_table"] = raw_data["data"]
        
        # Add preview
        structured["preview"] = {
            "has_balance_sheet": bool(structured["balance_sheet"]),
            "has_income_statement": bool(structured["income_statement"]),
            "has_cash_flow": bool(structured["cash_flow"])
        }
        
        return structured
    
    def _extract_balance_sheet(self, text: str) -> Dict[str, float]:
        """Extract balance sheet items from text"""
        bs = {}
        
        # Common balance sheet patterns
        patterns = {
            "total_assets": r"total\s+assets?\s*:?\s*(\d+[\d,\.]*)",
            "current_assets": r"current\s+assets?\s*:?\s*(\d+[\d,\.]*)",
            "total_liabilities": r"total\s+liabilities?\s*:?\s*(\d+[\d,\.]*)",
            "current_liabilities": r"current\s+liabilities?\s*:?\s*(\d+[\d,\.]*)",
            "equity": r"(?:shareholders?|stockholders?)\s+equity\s*:?\s*(\d+[\d,\.]*)",
            "cash": r"cash\s+(?:and\s+)?(?:cash\s+)?equivalents?\s*:?\s*(\d+[\d,\.]*)",
            "inventory": r"inventor(?:y|ies)\s*:?\s*(\d+[\d,\.]*)",
            "receivables": r"(?:accounts?\s+)?receivables?\s*:?\s*(\d+[\d,\.]*)"
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value_str = match.group(1).replace(',', '')
                try:
                    bs[key] = float(value_str)
                except ValueError:
                    pass
        
        return bs
    
    def _extract_income_statement(self, text: str) -> Dict[str, float]:
        """Extract income statement items from text"""
        is_data = {}
        
        patterns = {
            "revenue": r"(?:total\s+)?(?:revenue|sales)\s*:?\s*(\d+[\d,\.]*)",
            "gross_profit": r"gross\s+profit\s*:?\s*(\d+[\d,\.]*)",
            "operating_income": r"operating\s+income\s*:?\s*(\d+[\d,\.]*)",
            "net_income": r"net\s+(?:income|profit)\s*:?\s*(\d+[\d,\.]*)",
            "ebitda": r"ebitda\s*:?\s*(\d+[\d,\.]*)",
            "cogs": r"(?:cost\s+of\s+)?(?:goods\s+sold|revenue)\s*:?\s*(\d+[\d,\.]*)",
            "operating_expenses": r"operating\s+expenses?\s*:?\s*(\d+[\d,\.]*)"
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value_str = match.group(1).replace(',', '')
                try:
                    is_data[key] = float(value_str)
                except ValueError:
                    pass
        
        return is_data
    
    def _extract_cash_flow(self, text: str) -> Dict[str, float]:
        """Extract cash flow items from text"""
        cf = {}
        
        patterns = {
            "operating_cash_flow": r"(?:cash\s+from\s+)?operating\s+activities\s*:?\s*(\d+[\d,\.]*)",
            "investing_cash_flow": r"(?:cash\s+from\s+)?investing\s+activities\s*:?\s*(\d+[\d,\.]*)",
            "financing_cash_flow": r"(?:cash\s+from\s+)?financing\s+activities\s*:?\s*(\d+[\d,\.]*)",
            "free_cash_flow": r"free\s+cash\s+flow\s*:?\s*(\d+[\d,\.]*)"
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value_str = match.group(1).replace(',', '')
                try:
                    cf[key] = float(value_str)
                except ValueError:
                    pass
        
        return cf
